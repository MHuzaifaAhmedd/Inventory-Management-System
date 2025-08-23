"""
Advanced Reports Tab Module
Provides comprehensive reporting with charts and analytics
"""

import sys
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
                              QPushButton, QComboBox, QDateEdit, QTableWidget,
                              QTableWidgetItem, QHeaderView, QMessageBox,
                              QFileDialog, QTabWidget, QFrame, QGridLayout)
from PyQt5.QtCore import Qt, QDate, pyqtSignal
from PyQt5.QtGui import QFont, QColor
import datetime
from .currency_utils import format_currency, get_currency_symbol

try:
    import pandas as pd
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure
    import numpy as np
    from openpyxl import Workbook
    PANDAS_AVAILABLE = True
except ImportError as e:
    print(f"Warning: Advanced reporting dependencies not available: {e}")
    PANDAS_AVAILABLE = False

class AdvancedReportsTab(QWidget):
    """Advanced reports tab with charts and analytics"""
    
    def __init__(self, db_manager):
        super().__init__()
        self.db_manager = db_manager
        self.init_ui()
        
    def init_ui(self):
        """Initialize the user interface"""
        layout = QVBoxLayout(self)
        
        # Title
        title = QLabel("📊 Advanced Reports & Analytics")
        title.setFont(QFont("Segoe UI", 16, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        # Date range selection
        date_frame = QFrame()
        date_frame.setFrameStyle(QFrame.StyledPanel)
        date_layout = QHBoxLayout(date_frame)
        
        date_layout.addWidget(QLabel("From:"))
        self.date_from = QDateEdit()
        self.date_from.setDate(QDate.currentDate().addDays(-30))
        self.date_from.setCalendarPopup(True)
        date_layout.addWidget(self.date_from)
        
        date_layout.addWidget(QLabel("To:"))
        self.date_to = QDateEdit()
        self.date_to.setDate(QDate.currentDate())
        self.date_to.setCalendarPopup(True)
        date_layout.addWidget(self.date_to)
        
        self.refresh_btn = QPushButton("🔄 Refresh Reports")
        self.refresh_btn.clicked.connect(self.refresh_reports)
        date_layout.addWidget(self.refresh_btn)
        
        self.export_btn = QPushButton("📤 Export to Excel")
        self.export_btn.clicked.connect(self.export_data)
        date_layout.addWidget(self.export_btn)
        
        layout.addWidget(date_frame)
        
        # Tab widget for different report types
        self.tab_widget = QTabWidget()
        
        # Sales Report Tab
        self.sales_tab = self.create_sales_tab()
        self.tab_widget.addTab(self.sales_tab, "💰 Sales Report")
        
        # Product Performance Tab
        self.product_tab = self.create_product_tab()
        self.tab_widget.addTab(self.product_tab, "📦 Product Performance")
        
        # Profit Analysis Tab
        self.profit_tab = self.create_profit_tab()
        self.tab_widget.addTab(self.profit_tab, "💵 Profit Analysis")
        
        # Charts Tab
        if PANDAS_AVAILABLE:
            self.charts_tab = self.create_charts_tab()
            self.tab_widget.addTab(self.charts_tab, "📈 Charts & Analytics")
        
        layout.addWidget(self.tab_widget)
        
        # Initial load
        self.refresh_reports()
        
    def create_sales_tab(self):
        """Create sales report tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Summary cards
        summary_layout = QHBoxLayout()
        
        self.total_sales_label = QLabel(f"Total Sales: {get_currency_symbol()}0")
        self.total_sales_label.setFont(QFont("Segoe UI", 12, QFont.Bold))
        self.total_sales_label.setStyleSheet("background-color: #e8f5e8; padding: 10px; border-radius: 5px;")
        summary_layout.addWidget(self.total_sales_label)
        
        self.total_orders_label = QLabel("Total Orders: 0")
        self.total_orders_label.setFont(QFont("Segoe UI", 12, QFont.Bold))
        self.total_orders_label.setStyleSheet("background-color: #e8f5e8; padding: 10px; border-radius: 5px;")
        summary_layout.addWidget(self.total_orders_label)
        
        self.avg_order_label = QLabel(f"Avg Order: {get_currency_symbol()}0")
        self.avg_order_label.setFont(QFont("Segoe UI", 12, QFont.Bold))
        self.avg_order_label.setStyleSheet("background-color: #e8f5e8; padding: 10px; border-radius: 5px;")
        summary_layout.addWidget(self.avg_order_label)
        
        layout.addLayout(summary_layout)
        
        # Sales table
        self.sales_table = QTableWidget()
        self.sales_table.setColumnCount(6)
        self.sales_table.setHorizontalHeaderLabels([
            "Date", "Product", "Quantity", "Unit Price", "Total", "Profit"
        ])
        header = self.sales_table.horizontalHeader()
        for i in range(6):
            header.setSectionResizeMode(i, QHeaderView.Stretch)
        
        layout.addWidget(self.sales_table)
        
        return widget
        
    def create_product_tab(self):
        """Create product performance tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Product performance table
        self.product_table = QTableWidget()
        self.product_table.setColumnCount(7)
        self.product_table.setHorizontalHeaderLabels([
            "Product", "Total Sold", "Revenue", "Cost", "Profit", "Profit Margin %", "Stock Level"
        ])
        header = self.product_table.horizontalHeader()
        for i in range(7):
            header.setSectionResizeMode(i, QHeaderView.Stretch)
        
        layout.addWidget(self.product_table)
        
        return widget
        
    def create_profit_tab(self):
        """Create profit analysis tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Profit summary
        profit_summary_layout = QHBoxLayout()
        
        self.total_revenue_label = QLabel(f"Total Revenue: {get_currency_symbol()}0")
        self.total_revenue_label.setFont(QFont("Segoe UI", 12, QFont.Bold))
        self.total_revenue_label.setStyleSheet("background-color: #e8f5e8; padding: 10px; border-radius: 5px;")
        profit_summary_layout.addWidget(self.total_revenue_label)
        
        self.total_cost_label = QLabel(f"Total Cost: {get_currency_symbol()}0")
        self.total_cost_label.setFont(QFont("Segoe UI", 12, QFont.Bold))
        self.total_cost_label.setStyleSheet("background-color: #ffe8e8; padding: 10px; border-radius: 5px;")
        profit_summary_layout.addWidget(self.total_cost_label)
        
        self.total_profit_label = QLabel(f"Total Profit: {get_currency_symbol()}0")
        self.total_profit_label.setFont(QFont("Segoe UI", 12, QFont.Bold))
        self.total_profit_label.setStyleSheet("background-color: #e8f5e8; padding: 10px; border-radius: 5px;")
        profit_summary_layout.addWidget(self.total_profit_label)
        
        self.profit_margin_label = QLabel("Profit Margin: 0%")
        self.profit_margin_label.setFont(QFont("Segoe UI", 12, QFont.Bold))
        self.profit_margin_label.setStyleSheet("background-color: #e8f5e8; padding: 10px; border-radius: 5px;")
        profit_summary_layout.addWidget(self.profit_margin_label)
        
        layout.addLayout(profit_summary_layout)
        
        # Daily profit table
        self.daily_profit_table = QTableWidget()
        self.daily_profit_table.setColumnCount(4)
        self.daily_profit_table.setHorizontalHeaderLabels([
            "Date", "Revenue", "Cost", "Profit"
        ])
        header = self.daily_profit_table.horizontalHeader()
        for i in range(4):
            header.setSectionResizeMode(i, QHeaderView.Stretch)
        
        layout.addWidget(self.daily_profit_table)
        
        return widget
        
    def create_charts_tab(self):
        """Create charts and analytics tab"""
        if not PANDAS_AVAILABLE:
            widget = QWidget()
            layout = QVBoxLayout(widget)
            layout.addWidget(QLabel("Charts require pandas and matplotlib"))
            return widget
            
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Create matplotlib figure
        self.figure = Figure(figsize=(12, 8))
        self.canvas = FigureCanvas(self.figure)
        layout.addWidget(self.canvas)
        
        return widget
        
    def refresh_reports(self):
        """Refresh all reports with current data"""
        try:
            date_from = self.date_from.date().toString("yyyy-MM-dd")
            date_to = self.date_to.date().toString("yyyy-MM-dd")
            
            # Get sales data
            sales_data = self.db_manager.get_sales_report(date_from, date_to)
            
            # Update sales tab
            self.update_sales_tab(sales_data)
            
            # Update product performance tab
            self.update_product_tab(sales_data)
            
            # Update profit analysis tab
            self.update_profit_tab(sales_data)
            
            # Update charts if available
            if PANDAS_AVAILABLE:
                self.update_charts(sales_data)
                
        except Exception as e:
            print(f"Error refreshing reports: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.warning(self, "Error", f"Failed to refresh reports: {str(e)}\n\nPlease check the console for more details.")
            
    def update_sales_tab(self, sales_data):
        """Update sales report tab"""
        if not sales_data:
            self.sales_table.setRowCount(0)
            self.total_sales_label.setText("Total Sales: $0")
            self.total_orders_label.setText("Total Orders: 0")
            self.avg_order_label.setText("Avg Order: $0")
            return
            
        # Calculate summary
        total_sales = sum(row[4] for row in sales_data)  # total_amount
        total_orders = len(sales_data)
        avg_order = total_sales / total_orders if total_orders > 0 else 0
        
        self.total_sales_label.setText(f"Total Sales: {format_currency(total_sales)}")
        self.total_orders_label.setText(f"Total Orders: {total_orders}")
        self.avg_order_label.setText(f"Avg Order: {format_currency(avg_order)}")
        
        # Update table
        self.sales_table.setRowCount(len(sales_data))
        for row_idx, row in enumerate(sales_data):
            # Date, Product, Quantity, Unit Price, Total, Profit
            self.sales_table.setItem(row_idx, 0, QTableWidgetItem(str(row[5])[:10]))  # date
            self.sales_table.setItem(row_idx, 1, QTableWidgetItem(str(row[6])))  # product_name
            self.sales_table.setItem(row_idx, 2, QTableWidgetItem(str(row[2])))  # quantity
            self.sales_table.setItem(row_idx, 3, QTableWidgetItem(format_currency(row[3])))  # unit_price
            self.sales_table.setItem(row_idx, 4, QTableWidgetItem(format_currency(row[4])))  # total_amount
            self.sales_table.setItem(row_idx, 5, QTableWidgetItem(format_currency(row[4] - (row[2] * row[7]))))  # profit
            
    def update_product_tab(self, sales_data):
        """Update product performance tab"""
        if not sales_data:
            self.product_table.setRowCount(0)
            return
            
        # Group by product
        product_stats = {}
        for row in sales_data:
            product_name = row[6]
            quantity = row[2]
            unit_price = row[3]
            cost_price = row[7]
            
            if product_name not in product_stats:
                product_stats[product_name] = {
                    'total_sold': 0,
                    'revenue': 0,
                    'cost': 0,
                    'stock_level': 0
                }
            
            product_stats[product_name]['total_sold'] += quantity
            product_stats[product_name]['revenue'] += quantity * unit_price
            product_stats[product_name]['cost'] += quantity * cost_price
            
        # Get current stock levels
        for product_name in product_stats:
            product = self.db_manager.get_product_by_name(product_name)
            if product:
                product_stats[product_name]['stock_level'] = product['quantity']
        
        # Update table
        self.product_table.setRowCount(len(product_stats))
        for row_idx, (product_name, stats) in enumerate(product_stats.items()):
            profit = stats['revenue'] - stats['cost']
            profit_margin = (profit / stats['revenue'] * 100) if stats['revenue'] > 0 else 0
            
            self.product_table.setItem(row_idx, 0, QTableWidgetItem(product_name))
            self.product_table.setItem(row_idx, 1, QTableWidgetItem(str(stats['total_sold'])))
            self.product_table.setItem(row_idx, 2, QTableWidgetItem(format_currency(stats['revenue'])))
            self.product_table.setItem(row_idx, 3, QTableWidgetItem(format_currency(stats['cost'])))
            self.product_table.setItem(row_idx, 4, QTableWidgetItem(format_currency(profit)))
            self.product_table.setItem(row_idx, 5, QTableWidgetItem(f"{profit_margin:.1f}%"))
            self.product_table.setItem(row_idx, 6, QTableWidgetItem(str(stats['stock_level'])))
            
    def update_profit_tab(self, sales_data):
        """Update profit analysis tab"""
        if not sales_data:
            self.daily_profit_table.setRowCount(0)
            self.total_revenue_label.setText("Total Revenue: $0")
            self.total_cost_label.setText("Total Cost: $0")
            self.total_profit_label.setText("Total Profit: $0")
            self.profit_margin_label.setText("Profit Margin: 0%")
            return
            
        # Calculate totals
        total_revenue = sum(row[4] for row in sales_data)  # total_amount
        total_cost = sum(row[2] * row[7] for row in sales_data)  # quantity * cost_price
        total_profit = total_revenue - total_cost
        profit_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0
        
        self.total_revenue_label.setText(f"Total Revenue: {format_currency(total_revenue)}")
        self.total_cost_label.setText(f"Total Cost: {format_currency(total_cost)}")
        self.total_profit_label.setText(f"Total Profit: {format_currency(total_profit)}")
        self.profit_margin_label.setText(f"Profit Margin: {profit_margin:.1f}%")
        
        # Group by date
        daily_stats = {}
        for row in sales_data:
            date = str(row[5])[:10]  # date
            revenue = row[4]  # total_amount
            cost = row[2] * row[7]  # quantity * cost_price
            
            if date not in daily_stats:
                daily_stats[date] = {'revenue': 0, 'cost': 0}
            
            daily_stats[date]['revenue'] += revenue
            daily_stats[date]['cost'] += cost
        
        # Update daily profit table
        self.daily_profit_table.setRowCount(len(daily_stats))
        for row_idx, (date, stats) in enumerate(sorted(daily_stats.items())):
            profit = stats['revenue'] - stats['cost']
            
            self.daily_profit_table.setItem(row_idx, 0, QTableWidgetItem(date))
            self.daily_profit_table.setItem(row_idx, 1, QTableWidgetItem(format_currency(stats['revenue'])))
            self.daily_profit_table.setItem(row_idx, 2, QTableWidgetItem(format_currency(stats['cost'])))
            self.daily_profit_table.setItem(row_idx, 3, QTableWidgetItem(format_currency(profit)))
            
    def update_charts(self, sales_data):
        """Update charts with current data"""
        if not PANDAS_AVAILABLE or not sales_data:
            return
            
        try:
            # Clear previous charts
            self.figure.clear()
            
            # Create subplots
            ax1 = self.figure.add_subplot(2, 2, 1)  # Daily sales
            ax2 = self.figure.add_subplot(2, 2, 2)  # Product performance
            ax3 = self.figure.add_subplot(2, 2, 3)  # Profit trend
            ax4 = self.figure.add_subplot(2, 2, 4)  # Revenue vs Cost
            
            # Convert to pandas DataFrame
            df = pd.DataFrame(sales_data, columns=[
                'id', 'product_id', 'quantity', 'unit_price', 'total_amount', 
                'date', 'product_name', 'cost_price'
            ])
            df['date'] = pd.to_datetime(df['date'])
            df['profit'] = df['total_amount'] - (df['quantity'] * df['cost_price'])
            
            # Daily sales chart
            daily_sales = df.groupby(df['date'].dt.date)['total_amount'].sum()
            ax1.plot(daily_sales.index, daily_sales.values, marker='o')
            ax1.set_title('Daily Sales')
            ax1.set_xlabel('Date')
            ax1.set_ylabel(f'Sales ({get_currency_code()})')
            ax1.tick_params(axis='x', rotation=45)
            
            # Product performance chart
            product_sales = df.groupby('product_name')['total_amount'].sum().sort_values(ascending=True)
            ax2.barh(range(len(product_sales)), product_sales.values)
            ax2.set_yticks(range(len(product_sales)))
            ax2.set_yticklabels(product_sales.index)
            ax2.set_title('Product Sales Performance')
            ax2.set_xlabel(f'Sales ({get_currency_code()})')
            
            # Profit trend chart
            daily_profit = df.groupby(df['date'].dt.date)['profit'].sum()
            ax3.plot(daily_profit.index, daily_profit.values, marker='s', color='green')
            ax3.set_title('Daily Profit Trend')
            ax3.set_xlabel('Date')
            ax3.set_ylabel(f'Profit ({get_currency_code()})')
            ax3.tick_params(axis='x', rotation=45)
            
            # Revenue vs Cost pie chart
            total_revenue = df['total_amount'].sum()
            total_cost = (df['quantity'] * df['cost_price']).sum()
            total_profit = total_revenue - total_cost
            
            ax4.pie([total_cost, total_profit], labels=['Cost', 'Profit'], 
                   autopct='%1.1f%%', colors=['#ff9999', '#66b3ff'])
            ax4.set_title('Revenue Breakdown')
            
            self.figure.tight_layout()
            self.canvas.draw()
            
        except Exception as e:
            print(f"Error updating charts: {e}")
            
    def export_data(self):
        """Export report data to Excel"""
        try:
            filename, _ = QFileDialog.getSaveFileName(
                self, "Export Reports", "inventory_reports.xlsx", 
                "Excel Files (*.xlsx)"
            )
            
            if not filename:
                return
                
            # Create workbook
            wb = Workbook()
            
            # Sales Report Sheet
            ws1 = wb.active
            ws1.title = "Sales Report"
            ws1.append(["Date", "Product", "Quantity", f"Unit Price ({get_currency_code()})", f"Total ({get_currency_code()})", f"Profit ({get_currency_code()})"])
            
            for row in range(self.sales_table.rowCount()):
                row_data = []
                for col in range(self.sales_table.columnCount()):
                    item = self.sales_table.item(row, col)
                    row_data.append(item.text() if item else "")
                ws1.append(row_data)
            
            # Product Performance Sheet
            ws2 = wb.create_sheet("Product Performance")
            ws2.append(["Product", "Total Sold", f"Revenue ({get_currency_code()})", f"Cost ({get_currency_code()})", f"Profit ({get_currency_code()})", "Profit Margin %", "Stock Level"])
            
            for row in range(self.product_table.rowCount()):
                row_data = []
                for col in range(self.product_table.columnCount()):
                    item = self.product_table.item(row, col)
                    row_data.append(item.text() if item else "")
                ws2.append(row_data)
            
            # Profit Analysis Sheet
            ws3 = wb.create_sheet("Profit Analysis")
            ws3.append(["Date", f"Revenue ({get_currency_code()})", f"Cost ({get_currency_code()})", f"Profit ({get_currency_code()})"])
            
            for row in range(self.daily_profit_table.rowCount()):
                row_data = []
                for col in range(self.daily_profit_table.columnCount()):
                    item = self.daily_profit_table.item(row, col)
                    row_data.append(item.text() if item else "")
                ws3.append(row_data)
            
            # Save file
            wb.save(filename)
            QMessageBox.information(self, "Export Successful", 
                                  f"Reports exported to: {filename}")
            
        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"Failed to export: {str(e)}")
